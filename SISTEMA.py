import tkinter as tk
from tkinter import messagebox
import datetime
from tkcalendar import Calendar
import openpyxl
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
import os
import pandas as pd

# MENU PRINCIPAL
def criar_janela_principal():
    root = tk.Tk()
    root.title("BACK SISTEMAS ")
    root.geometry("1000x690")
    root.configure(bg="black")

    # Função para cadastrar despesas
    def cadastrar_despesas():
        root.withdraw()  # Oculta a janela principal
        nova_janela = tk.Toplevel(root)
        nova_janela.title("Cadastrar Despesas")
        nova_janela.geometry("1250x750")
        nova_janela.configure(bg="#222222")
        nova_janela.resizable(True, True)
        nova_janela.maxsize(width=1400, height=900)
        nova_janela.minsize(width=500, height=400)

        # abrir powerbi
        def abrir_powerbi():
            caminho_powerbi = "despesas.pbix"
            try:
                os.startfile(caminho_powerbi)
            except FileNotFoundError:
                print("Arquivo Power BI não encontrado.")
            except OSError:
                print("Erro ao abrir o arquivo.")

        def limpar_campos():
            entrada_descricao.delete(0, tk.END)
            entrada_quantidade.delete(0, tk.END)
            entrada_valor.delete(0, tk.END)
            entrada_parcelas.delete(0, tk.END)
            entry_dataVencimento.delete(0, tk.END)

        def gerar_txt_despesas():
            data = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            descriao = entrada_descricao.get()
            quantidade = entrada_quantidade.get()
            valor_uni = entrada_valor.get()
            parcelas = entrada_parcelas.get()
            data_vencimento = entry_dataVencimento.get()
            valor_total = float(quantidade) * float(valor_uni)
            valor_parcela = float(valor_total) / float(parcelas)

            with open('despesas.txt', 'a') as arquivo:
                arquivo.write('\n')
                arquivo.write("BACK INFORMÁTICA\n")
                arquivo.write("soluções tecnologicas\n")
                arquivo.write("fone: 91 983252639\n")
                arquivo.write("------------------------------------\n")
                arquivo.write(f"Data da transação: {data}\n")
                arquivo.write(f"Descrição: {descriao}\n")
                arquivo.write(f"Quantidade: {quantidade}\n")
                arquivo.write(f"Valor unitário: {valor_uni}\n")
                arquivo.write("------------------------------------\n")
                arquivo.write(f"Parcelas: {parcelas}\n")
                arquivo.write(f"Data de vencimento: {data_vencimento}\n")
                arquivo.write(f"Valor total: {valor_total}\n")
                arquivo.write(f"Valor parcela: {valor_parcela}\n")
                arquivo.write("\n")
                arquivo.write("**************************************")

            messagebox.showinfo("Sucesso", "Arquivo gerado com sucesso!")

        def salvar_dados_despesa():
            data_entrada = datetime.datetime.now().strftime("%Y-%m-%d")
            descricao = entrada_descricao.get()
            quantidade = int(entrada_quantidade.get())
            valorUN = float(entrada_valor.get())
            parcelas = int(entrada_parcelas.get())
            data_vencimento = entry_dataVencimento.get()
            valor_total = float(valorUN) * float(quantidade)
            valor_parcela = float(valor_total) / float(parcelas)
            try:
                # Verifica se o arquivo já existe
                workbook = load_workbook('despesas.xlsx')
                sheet = workbook.active

            except FileNotFoundError:
                # Cria um novo arquivo se não existir
                workbook = Workbook()
                sheet = workbook.active
                sheet.append(['data_entrada', 'descrição', 'quantidade', 'valorUN', 'parcelas', 'Data_vencimento','valor_total','valor_parcela'])

            # Adiciona uma nova linha com os INSTITUICAO_ENSINO
            sheet.append([data_entrada, descricao, quantidade, valorUN, parcelas, data_vencimento,valor_total,valor_parcela])
            workbook.save('despesas.xlsx')

            mensagem = (tk.messagebox.showinfo("Beleza!", "Despesa cadastrada, clique OK para adicionar outra despesa!"))

            gerar_txt_despesas()
            limpar_campos()

        def buscar_data():
            data_selecionada = cal.get_date()
            entry_dataVencimento.delete(0, tk.END)
            entry_dataVencimento.insert(0, data_selecionada)

        #_________________________BANNER DA JANELA PRINCIPAL_______________________________
        banner = tk.Label(nova_janela, text="Cadastro de Despesas", bg="#222222", fg="#00BFFF", font=("System", 25, "bold"))
        banner.place(x=20, y=10)
        # MOSTRAR A DATA E HORA ATUAL
        label_data_entrada = tk.Label(nova_janela, text=datetime.datetime.now().strftime("%d/%m/%Y %H:%M:%S"),bg="#222222", fg="white", font=("System", 10, "bold"))
        label_data_entrada.place(x=25, y=55)
        #_______________________PRIMEIRA LINHA DO FORMULARIO____________________________________
        # formulario despesas
        label_quantidade = tk.Label(nova_janela, text="Qnt ", bg="#222222", fg="white", font=("consolas", 12, "bold"))
        label_quantidade.place(x=20, y=105)
        entrada_quantidade = tk.Entry(nova_janela, bg="black", bd=0, fg="white", font=("consolas", 12, "bold"))
        entrada_quantidade.place(x=20, y=130, width=50, height=40)

        label_descricao = tk.Label(nova_janela, text="Descrição ", bg="#222222", fg="white", font=("consolas", 12, "bold"))
        label_descricao.place(x=95, y=105)
        entrada_descricao = tk.Entry(nova_janela, bg="black", bd=0, fg="white", font=("consolas", 12, "bold"))
        entrada_descricao.place(x=95, y=130, width=500, height=40)
        # _______________________SEGUNDA LINHA DO FORMULARIO____________________________________
        label_valor = tk.Label(nova_janela, text="Valor", bg="#222222", fg="white", font=("consolas", 12, "bold"))
        label_valor.place(x=20, y=190)
        entrada_valor = tk.Entry(nova_janela, bg="black", bd=0, fg="white", font=("consolas", 12, "bold"))
        entrada_valor.place(x=20, y=215, width=100, height=40)

        label_parcelas = tk.Label(nova_janela, text="Parcelas", bg="#222222", fg="white",font=("consolas", 12, "bold"))
        label_parcelas.place(x=140, y=190)
        entrada_parcelas = tk.Entry(nova_janela, bg="black", bd=0, fg="white", font=("consolas", 12, "bold"))
        entrada_parcelas.place(x=140, y=215, width=150, height=40)

        # entrada dos INSTITUICAO_ENSINO de vencimento
        label_data_vencimento = tk.Label(nova_janela, text="data_vencimento", bg="#222222", fg="white",font=("consolas", 12, "bold"))
        label_data_vencimento.place(x=310, y=190)
        entry_dataVencimento = tk.Entry(nova_janela, bg="black", bd=0, fg="white", font=("consolas", 12, "bold"))
        entry_dataVencimento.place(x=310, y=215, width=160, height=40)

        # Botão para salvar a despesa
        botao_salvar = tk.Button(nova_janela, text="Salvar Despesa", command=salvar_dados_despesa, bg="white",fg="#222222", font=("consolas", 12, "bold"))
        botao_salvar.place(x=20, y=290, width=400, height=50)

        # calendário com botão para selecionar a data
        label_info = tk.Label(nova_janela, text="Selecione a Data de Vencimento", bg="#222222", fg="white",font=("consolas", 12, "bold"))
        label_info.place(x=710, y=70)
        cal = Calendar(nova_janela, selectmode='day', year=2023, month=11, day=22)
        cal.place(x=730, y=100)
        select_data_vencimento = tk.Button(nova_janela, text="Selecionar", command=buscar_data, bg="#444444", bd=0,fg="#00BFFF", font=("consolas", 12, "bold"))
        select_data_vencimento.place(x=790, y=310, width=150, height=50)

        # FRAME PARA ALOCAR A TABELA
        frame = tk.Frame(nova_janela, bg="#222222")
        frame.place(x=20, y=370, width=1220, height=280)

        df = pd.read_excel('despesas.xlsx')
        texto_tabela = tk.Text(frame, bg="Black",bd=0, fg="#00BFFF", font=("consolas", 12, "bold"))
        texto_tabela.pack(fill="both", expand=True)
        texto_tabela.insert("end", df.to_string())

        botao_voltar = tk.Button(nova_janela, text="Voltar", font=("consolas", 12), bg="#444444", bd=0, fg="#00BFFF",command=lambda: voltar_para_janela_principal(nova_janela))
        botao_voltar.place(x=20, y=660, width=400, height=50)

        botao_analise_gastos = tk.Button(nova_janela, text="Analisar Gastos", bg="#444444", bd=0, fg="#00BFFF",font=("consolas", 12, "bold"), command=abrir_powerbi)
        botao_analise_gastos.place(x=430, y=660, width=400, height=50)

        botao_view = tk.Button(nova_janela, text="Visualizar Tabela", bg="#444444", bd=0, fg="#00BFFF",font=("consolas", 12, "bold"), command=None)
        botao_view.place(x=840, y=660, width=400, height=50)

    # Função cadastrar vendas
    def cadastrar_vendas():
        root.withdraw()  # Oculta a janela principal
        nova_janela = tk.Toplevel(root)
        nova_janela.title("Cadastrar Vendas")
        nova_janela.geometry("1250x750")
        nova_janela.configure(bg="#222222")
        nova_janela.resizable(True, True)
        nova_janela.maxsize(width=1300, height=800)
        nova_janela.minsize(width=500, height=400)

        # abrir powerbi
        def abrir_powerbi():
            caminho_powerbi = "vendas.pbix"
            try:
                os.startfile(caminho_powerbi)
            except FileNotFoundError:
                print("Arquivo Power BI não encontrado.")
            except OSError:
                print("Erro ao abrir o arquivo.")
        # Função para limpar os campos
        def limpar_campos():
            entrada_item.delete(0, tk.END)
            entrada_descricao.delete(0, tk.END)
            entrada_nova_quantidade.delete(0, tk.END)
            entrada_valor_uni.delete(0, tk.END)
            entrada_valor_pago.delete(0, tk.END)

        # Função para gerar o arquivo txt
        def gerar_txt():
            data = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            item = entrada_item.get()
            descricao = entrada_descricao.get()
            quantidade = entrada_nova_quantidade.get()
            valor_uni = entrada_valor_uni.get()
            valor_pago = entrada_valor_pago.get()
            valor_total = float(valor_uni) * float(quantidade)
            troco = float(valor_pago) - float(valor_total)

            with open("vendas.txt", "a") as f:
                f.write("\n")
                f.write(f"BACK INFORMATICA\n")
                f.write(f"soluções tecnologicas\n")
                f.write(f"fone: 91 983252639\n")
                f.write(f"------------------------------------\n")
                f.write(f"Data da transação: {data}\n")
                f.write(f"Item: {item}\n")
                f.write(f"Descrição: {descricao}\n")
                f.write(f"Quantidade: {quantidade}\n")
                f.write(f"Valor unitário: {valor_uni}\n")
                f.write(f"------------------------------------\n")
                f.write(f"Valor pago: {valor_pago}\n")
                f.write(f"Total: {valor_total}\n")
                f.write(f"Troco: {troco}\n")
                f.write("\n")
                f.write(f"Muito obrigado por usar nosso sistema\n")
                f.write("\n")
                f.write(f"* * * * * * * * * * * * * * * * * * * *")
            messagebox.showinfo("Sucesso", "Arquivo gerado com sucesso!")

        # Funções para interagir com o Excel
        def salvar_vendas():
            nova_data = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            novo_item = entrada_item.get()
            nova_descricao = entrada_descricao.get()
            nova_quantidade = entrada_nova_quantidade.get()
            novo_valor_uni = entrada_valor_uni.get()
            novo_valor_pago = entrada_valor_pago.get()
            novo_valor_total = float(novo_valor_uni) * float(nova_quantidade)
            troco = float(novo_valor_pago) - float(novo_valor_total)

            try:
                # Verifica se o arquivo já existe
                workbook = load_workbook('vendas.xlsx')
                sheet = workbook.active

            except FileNotFoundError:
                # Cria um novo arquivo se não existir
                workbook = Workbook()
                sheet = workbook.active
                sheet.append(
                    ['Data da transação', 'Item', 'Descrição', 'Quantidade', 'Valor unitário', 'Valor pago', 'Total',
                     'Troco'])

            # Adiciona uma nova linha com os INSTITUICAO_ENSINO
            sheet.append([nova_data, novo_item, nova_descricao, nova_quantidade, novo_valor_uni, novo_valor_pago,
                          novo_valor_total, troco])
            workbook.save('vendas.xlsx')

            # Exibir a mensagem de confirmação
            mensagem = (tk.messagebox.showinfo("OK!", "Despesa cadastrada, clique OK para adicionar outra despesa."))
            gerar_txt()
            limpar_campos()

        # CRIANDO O FORMULARIO COM OS CAMPOS NECESSARIOS PARA O CADASTRO DE VENDAS

        # _________________________BANNER DA JANELA PRINCIPAL_______________________________
        banner = tk.Label(nova_janela, text="Cadastro de Vendas", bg="#222222", fg="#00BFFF",font=("System", 25, "bold"))
        banner.place(x=20, y=10)

        # MOSTRAR A DATA E HORA ATUAL
        label_data_entrada = tk.Label(nova_janela, text=datetime.datetime.now().strftime("%d/%m/%Y %H:%M:%S"),bg="#222222", fg="white", font=("System", 10, "bold"))
        label_data_entrada.place(x=25, y=55)
        #____________________________PRIMEIRA LINHA DO FORMULARIO____________________________
        label_item = tk.Label(nova_janela, text="Item", font=("consolas", 12), bg="#222222", fg="white");label_item.place(x=20, y=100, width=50, height=40)
        entrada_item = tk.Entry(nova_janela, font=("consolas", 12), bg="black", bd=0, fg="white");entrada_item.place(x=20, y=130, width=150, height=40)

        label_descri = tk.Label(nova_janela, text="Descrição", font=("consolas", 12), bg="#222222", fg="white");label_descri.place(x=190, y=100, width=120, height=40)
        entrada_descricao = tk.Entry(nova_janela, font=("consolas", 12), bg="black", bd=0, fg="white");entrada_descricao.place(x=190, y=130, width=400, height=40)
        #_____________________________SEGUNDA LINHA DO FORMULARIO____________________________
        labe_quant = tk.Label(nova_janela, text="Qtd.: ", font=("consolas", 12), bg="#222222", fg="white");labe_quant.place(x=20, y=180, width=80, height=40)
        entrada_nova_quantidade = tk.Entry(nova_janela, font=("consolas", 12), bg="black", bd=0, fg="white");entrada_nova_quantidade.place(x=20, y=210, width=80, height=40)

        label_valor_uni = tk.Label(nova_janela, text="ValorUnitário:", font=("consolas", 12), bg="#222222", fg="white");label_valor_uni.place(x=130, y=180, width=160, height=40)
        entrada_valor_uni = tk.Entry(nova_janela, font=("consolas", 12), bg="black", bd=0, fg="white");entrada_valor_uni.place(x=130, y=210, width=160, height=40)

        label_valor_pago = tk.Label(nova_janela, text="valorPago:", font=("consolas", 12), bg="#222222", fg="white");label_valor_pago.place(x=310, y=180, width=150, height=40)
        entrada_valor_pago = tk.Entry(nova_janela, font=("consolas", 12), bg="black", bd=0, fg="white");entrada_valor_pago.place(x=320, y=210, width=150, height=40)

        #BOTÃO PARA SALVAR A VENDA
        botao_inserir = tk.Button(nova_janela, text="Salvar venda",  font=("consolas", 12), bg="white", bd=0, fg="#222222", command=salvar_vendas)
        botao_inserir.place(x=20, y=280, width=300, height=40)

        botao_voltar = tk.Button(nova_janela, text="Voltar", font=("consolas", 12), bg="#444444", bd=0, fg="#00BFFF",command=lambda: voltar_para_janela_principal(nova_janela))
        botao_voltar.place(x=20, y=620, width=400, height=50)

        botao_analise = tk.Button(nova_janela, text="Análise de Vendas", font=("consolas", 12), bg="#444444", bd=0, fg="#00BFFF",command=None)
        botao_analise.place(x=430, y=620, width=400, height=50)

        botao_view = tk.Button(nova_janela, text="Visualizar Tabela", font=("consolas", 12), bg="#444444", bd=0, fg="#00BFFF",command=None )
        botao_view.place(x=840, y=620, width=400, height=50)

        frame = tk.Frame(nova_janela, bg="#222222")
        frame.place(x=20, y=350, width=1220, height=250)
        df = pd.read_excel('vendas.xlsx')
        texto_tabela = tk.Text(frame, bg="BLACK", fg="#00BFFF", font=("consolas", 12, "bold"))
        texto_tabela.pack(fill="both", expand=True)
        texto_tabela.insert("end", df.to_string())

    def add_estoque():
        root.withdraw()  # Oculta a janela principal
        nova_janela = tk.Toplevel(root)
        nova_janela.title("Cadastrar Estoque")
        nova_janela.geometry("1250x750")
        nova_janela.configure(bg="#222222")
        nova_janela.resizable(True, True)
        nova_janela.maxsize(width=1900, height=800)
        nova_janela.minsize(width=500, height=400)

        def abrir_powerbi():
            # Substitua pelo caminho do seu arquivo Power BI
            caminho_powerbi = "estoque.pbix"

            # Abrir o arquivo Power BI
            try:
                os.startfile(caminho_powerbi)
            except FileNotFoundError:
                print("Arquivo Power BI não encontrado.")
            except OSError:
                print("Erro ao abrir o arquivo.")

        def limpar_campos():
            codigo_barras_entry.delete(0, tk.END)
            descricao_entry.delete(0, tk.END)
            quantidade_entry.delete(0, tk.END)
            valor_inicial_entry.delete(0, tk.END)
            valor_final_entry.delete(0, tk.END)

        def salvar_estoque():
            codigo_barras = codigo_barras_entry.get()
            descricao = descricao_entry.get()
            quantidade = quantidade_entry.get()
            valor_inicial = valor_inicial_entry.get()
            valor_final = valor_final_entry.get()
            lucro = float(valor_final) - float(valor_inicial)
            lucroTotal = float(valor_final) * float(quantidade)

            try:
                workbook = openpyxl.load_workbook('estoque.xlsx')
                codigo_barras = codigo_barras_entry.get()
                descricao = descricao_entry.get()
                quantidade = quantidade_entry.get()
                valor_inicial = valor_inicial_entry.get()
                valor_final = valor_final_entry.get()
                lucro = float(valor_final) - float(valor_inicial)
                lucroTotal = float(valor_final) * float(quantidade)
                sheet = workbook.active
            except FileNotFoundError:
                workbook = openpyxl.Workbook()
                sheet = workbook.active
                sheet.append(['Código de Barras', 'Descrição', 'Quantidade', 'Valor Inicial', 'Valor Final', 'Lucro',
                              'lucroTotal'])

            # Adicionar uma nova linha com os INSTITUICAO_ENSINO do produto
            sheet.append([codigo_barras, descricao, quantidade, valor_inicial, valor_final, lucro, lucroTotal])

            # Salvar o arquivo
            workbook.save('estoque.xlsx')

            limpar_campos()

        # _________________________BANNER DA JANELA PRINCIPAL_______________________________
        banner = tk.Label(nova_janela, text="Cadastro de Estoque", bg="#222222", fg="#00BFFF",font=("System", 25, "bold"))
        banner.place(x=20, y=10)
        # MOSTRAR A DATA E HORA ATUAL
        label_data_entrada = tk.Label(nova_janela, text=datetime.datetime.now().strftime("%d/%m/%Y %H:%M:%S"),bg="#222222", fg="white", font=("System", 10, "bold"))
        label_data_entrada.place(x=25, y=55)

        label_codigo = tk.Label(nova_janela, text="Cod.", fg="white", bg="#222222", font=("consolas", 12, "bold"));
        label_codigo.place(x=20, y=110)
        codigo_barras_entry = tk.Entry(nova_janela, font=("consolas", 12), bg="black", bd=0, fg="white");
        codigo_barras_entry.place(x=20, y=135, width=60, height=40)

        label_quantidade = tk.Label(nova_janela, text="Qtd.", fg="white", bg="#222222", font=("consolas", 12, "bold"))
        label_quantidade.place(x=100, y=110)
        quantidade_entry = tk.Entry(nova_janela, font=("consolas", 12), bg="black", bd=0, fg="white")
        quantidade_entry.place(x=100, y=135, width=60, height=40)

        label_descricao = tk.Label(nova_janela, text="Desc.", fg="white", bg="#222222",font=("consolas", 12, "bold"));
        label_descricao.place(x=185, y=110)
        descricao_entry = tk.Entry(nova_janela, font=("consolas", 12), bg="black", bd=0, fg="white")
        descricao_entry.place(x=185, y=135, width=300, height=40)

        label_valor_uni = tk.Label(nova_janela, text="valorInicial", fg="white", bg="#222222",font=("consolas", 12, "bold"))
        label_valor_uni.place(x=20, y=200)
        valor_inicial_entry = tk.Entry(nova_janela, font=("consolas", 12), bg="black", bd=0, fg="white")
        valor_inicial_entry.place(x=20, y=225, width=150, height=40)

        label_valor_final = tk.Label(nova_janela, text="Valor_final", fg="white", bg="#222222",font=("consolas", 12, "bold"))
        label_valor_final.place(x=190, y=200)
        valor_final_entry = tk.Entry(nova_janela, font=("consolas", 12), bg="black", bd=0, fg="white")
        valor_final_entry.place(x=190, y=225, width=150, height=40)

        botao_inserir = tk.Button(nova_janela, text="Salvar venda", font=("consolas", 12), bg="white", bd=0,fg="#222222", command=salvar_estoque)
        botao_inserir.place(x=20, y=300, width=400, height=40)

        frame = tk.Frame(nova_janela, bg="#222222")
        frame.place(x=20, y=380, width=1220, height=250)
        df = pd.read_excel('estoque.xlsx')
        texto_tabela = tk.Text(frame, bg="BLACK", fg="#00BFFF", font=("consolas", 12, "bold"))
        texto_tabela.pack(fill="both", expand=True)
        texto_tabela.insert("end", df.to_string())

        botao_voltar = tk.Button(nova_janela, text="Voltar", font=("consolas", 12), bg="#444444", bd=0, fg="#00BFFF",command=lambda: voltar_para_janela_principal(nova_janela))
        botao_voltar.place(x=20, y=650, width=400, height=50)

        botao_analise = tk.Button(nova_janela, text="Análisa Estoque", font=("consolas", 12), bg="#444444", bd=0, fg="#00BFFF",command=None)
        botao_analise.place(x=430, y=650, width=400, height=50)

        botao_view = tk.Button(nova_janela, text="Visualizar Tabela", font=("consolas", 12), bg="#444444", bd=0, fg="#00BFFF",command=None)
        botao_view.place(x=840, y=650, width=400, height=50)

    #RELATORIOS
    def relatorios():
        root.withdraw()  # Oculta a janela principal
        nova_janela = tk.Toplevel(root)
        nova_janela.title("Relatórios")
        nova_janela.geometry("1200x700")
        nova_janela.configure(bg="black")


        # _________________________BANNER DA JANELA PRINCIPAL_______________________________
        banner = tk.Label(nova_janela, text="Analise e relatorios do sistema", bg="black", fg="#00BFFF",font=("System", 25, "bold"))
        banner.place(x=20, y=10)

        # MOSTRAR A DATA E HORA ATUAL
        label_data_entrada = tk.Label(nova_janela, text=datetime.datetime.now().strftime("%d/%m/%Y   %H:%M:%S"), bg="black", fg="white", font=("System", 12, "bold"))
        label_data_entrada.place(x=120, y=60)

        frame_grafico_vendas = tk.Frame(nova_janela, bg="#222222")
        frame_grafico_vendas.place(x=400, y=100, width=250, height=480)
        label_vendas_dash = tk.Label(frame_grafico_vendas, text="Vendas", font="consolas 12 bold", bg="#222222", fg="#00bfff")
        label_vendas_dash.place(x=10,y=10)

        frame_grafico_despesas = tk.Frame(nova_janela, bg="#222222")
        frame_grafico_despesas.place(x=665, y=100, width=250, height=480)
        label_despesa_dash = tk.Label(frame_grafico_despesas, text="Despesas", font="consolas 12 bold", bg="#222222", fg="#00bfff")
        label_despesa_dash.place(x=10, y=10)

        frame_grafico_estoque = tk.Frame(nova_janela, bg="#222222")
        frame_grafico_estoque.place(x=930, y=100, width=250, height=480)
        label_estoque_dash = tk.Label(frame_grafico_estoque, text="Estoque", font="consolas 12 bold", bg="#222222", fg= "#00bfff")
        label_estoque_dash.place(x=10, y=10)

        botao_voltar = tk.Button(nova_janela, text="Voltar",font=("consolas", 12, "bold"), fg="#222222", bg="#00BFFF", bd=0,command=lambda: voltar_para_janela_principal(nova_janela))
        botao_voltar.place(x=20, y=600, width=360, height=50)

    #cadastre aqui os USUARIOS
    def usuarios():
        root.withdraw()  # Oculta a janela principal
        nova_janela = tk.Toplevel(root)
        nova_janela.title("Usuarios")
        nova_janela.configure(bg="#222222")
        nova_janela.resizable(True, True)
        nova_janela.maxsize(width=1300, height=700)
        nova_janela.minsize(width=900, height=700)

        # _________________________BANNER DA JANELA PRINCIPAL_______________________________
        banner = tk.Label(nova_janela, text="Usuarios", bg="#222222", fg="#00BFFF",font=("System", 25, "bold"))
        banner.place(x=20, y=10)

        # MOSTRAR A DATA E HORA ATUAL
        label_data_entrada = tk.Label(nova_janela, text=datetime.datetime.now().strftime("%d/%m/%Y %H:%M:%S"),bg="#222222", fg="white", font=("System", 10, "bold"))
        label_data_entrada.place(x=25, y=55)

        # Função para cadastrar um usuário

        def limpar_campos():
            entry_user.delete(0, tk.END)
            entry_senha.delete(0, tk.END)
        def cadastrar_usuario():
            user = entry_user.get()
            senha = entry_senha.get()

            try:
                workbook = openpyxl.load_workbook('usuarios.xlsx')
                sheet = workbook.active
                sheet.append([user, senha])
                workbook.save('usuarios.xlsx')
                messagebox.showinfo("Sucesso", "Usuário cadastrado com sucesso!")
            except Exception as e:
                workbook = openpyxl.Workbook()
                sheet = workbook.active
                sheet.append([user, senha])
                workbook.save('usuarios.xlsx')
                messagebox.showerror("Erro", f"Erro ao cadastrar usuário: {str(e)}")
            limpar_campos()

        label_user = tk.Label(nova_janela, text="Usuário:", bg="#222222", fg="white",font=("System", 12, "bold"))
        label_user.place(x=20, y=100)
        entry_user = tk.Entry(nova_janela, font=("consolas", 12), bg="black", bd=0, fg="white")
        entry_user.place(x=20, y=130)

        label_senha = tk.Label(nova_janela, text="Senha:", bg="#222222", fg="white",font=("System", 12, "bold"))
        label_senha.place(x=20, y=160)
        entry_senha = tk.Entry(nova_janela, show="*", font=("consolas", 12), bg="black", bd=0, fg="white")
        entry_senha.place(x=20, y=190)

        botão_cadastrar_usuario = tk.Button(nova_janela, text="Cadastrar",font=("consolas", 12, "bold"), fg="#222222", bg="#00BFFF", bd=0,command=cadastrar_usuario)
        botão_cadastrar_usuario.place(x=20, y=220, width=200, height=50)
        # Função para editar um usuário
        def editar_usuario():
            pass

        # ... (lógica para buscar o usuário e permitir a edição)

        # Função para remover um usuário
        def remover_usuario():
            pass

        botao_editar = tk.Button(nova_janela, text="Editar",font=("consolas", 12, "bold"), fg="#222222", bg="#00BFFF", bd=0,command=editar_usuario)
        botao_editar.place(relx=0.5, rely=0.4, anchor="center", width=200, height=50)

        botao_remover = tk.Button(nova_janela, text="Remover",font=("consolas", 12, "bold"), fg="#222222", bg="#00BFFF", bd=0,command=remover_usuario)
        botao_remover.place(relx=0.5, rely=0.5, anchor="center", width=200, height=50)

        botao_voltar = tk.Button(nova_janela, text="Voltar",font=("consolas", 12, "bold"), fg="#222222", bg="#00BFFF", bd=0,command=lambda: voltar_para_janela_principal(nova_janela))
        botao_voltar.place(relx=0.5, rely=0.9, anchor="center", width=100, height=50)

    #informações do programa
    def informa():
        root.withdraw()  # Oculta a janela principal
        nova_janela = tk.Toplevel(root)
        nova_janela.title("informações")
        nova_janela.configure(bg="#222222")
        nova_janela.resizable(True, True)
        nova_janela.maxsize(width=1300, height=700)
        nova_janela.minsize(width=900, height=700)

        banner = tk.Label(nova_janela, text="BACK INFORMÁTICA\n"
                                            "Soluções tecnologicas\n"
                                            "\n"
                                            "* Esta versão 1,0 feita para cadastrar despesas,vendas e estoque em arquivos simples como \n"
                                            "o excel e para analise de INSTITUICAO_ENSINO com power bi, atravez de uma interface amigavel com temas\n"
                                            "escuro, para melhor visibilidade e conforto \n"
                                            " \n"
                                            "\n"
                                            "* desenvolvido por: AIRTON JUNIOR\n"
                                            "TECNOLOGO E ESTUDANTE DO CURSO DE "
                                            "ANANLISE E DESENVOLVIMENTO DE SOFTWARES\n"
                                            "\n"
                                            "\n"
                                            " Obrigado por usar meu sistema\n"
                                    , bg="#222222", fg="red", font=("consolas", 12, "bold"))
        banner.pack()

        botao_voltar = tk.Button(nova_janela, text="Voltar", fg="white", bg="#333333", bd=0, command=lambda: voltar_para_janela_principal(nova_janela))
        botao_voltar.place(relx=0.5, rely=0.9, anchor="center", width=100, height=50)

    # Função para enviar mensagem no whatsapp
    def send_whatsapp_message():
        phone_number = "+5591983252639"
        message = "ola, pertenço ao software da serie v1.0, aguardando retorno, obrigado !"
        kit.sendwhatmsg_instantly(phone_number, message)

    #ESTAS FUNÇÓES ESTÃO EM VARIAS JANELAS DA INTERFACE, FUNDAMENTAIS.
    def abrir_nova_janela(nome_botao):
        root.withdraw()  # Oculta a janela principal
        nova_janela = tk.Toplevel(root)
        nova_janela.title(nome_botao)

        label = tk.Label(nova_janela, text=f"Você clicou no {nome_botao}")
        label.pack(pady=10)

        botao_voltar = tk.Button(nova_janela, text="Voltar", command=lambda: voltar_para_janela_principal(nova_janela))
        botao_voltar.pack(pady=10)

    # Função para voltar para a janela principal
    def voltar_para_janela_principal(janela_atual):
        janela_atual.destroy()
        root.deiconify()  # Mostra a janela principal novamente

    # função para fefchar janela principal
    def fechar():
        root.destroy()

    #INTERFACE PRINCIPAL
    frame_logo = tk.Frame(root, bg="black", bd=0)
    frame_logo.place(x=330, y=10, width=660, height=670)

    label_logo = tk.Label(frame_logo, text="Seu logo", bg="black", fg="#00BFFF", font=("consolas", 25, "bold"))
    label_logo.place(x=0, y=0, width=660, height=670)

    hora = tk.Label(frame_logo, text="v1.0", bg="black", fg="#00BFFF", font=("consolas", 12, "bold"))
    hora.place(x=570, y=645)

    tk.Button(root, text="Despesas", fg="#00BFFF", bg="#222222", border=0, font=("consolas", 10, "bold"), command= cadastrar_despesas).place(x=10, y=10, width=150, height=300)
    tk.Button(root, text="Vendas", fg="#00BFFF", bg="#222222", border=0, font=("consolas", 10, "bold"), command= cadastrar_vendas).place(x=10, y=380, width=150, height=300)
    tk.Button(root, text="Estoque", fg="#00BFFF", bg="#222222", border=0, font=("consolas", 10, "bold"), command= add_estoque).place(x=170, y=200, width=150, height=300)
    tk.Button(root, text=" Informações ", fg="#222222", bg="#00BFFF", border=0, font=("consolas", 10, "bold"), command= informa).place(x=170, y=140, width=150, height=50)
    tk.Button(root, text=" Sair/Fechar ", fg="#222222", bg="#00BFFF", border=0, font=("consolas", 10, "bold"), command= fechar).place(x=170, y=510, width=150, height=50)
    tk.Button(root, text="Suporte", fg="#222222", bg="#00BFFF", border=0, font=("consolas", 10, "bold"), command= send_whatsapp_message).place(x=10, y=320, width=150, height=50)
    tk.Button(root, text="Relatorios", fg="#00BFFF", bg="#222222", border=0, font=("consolas", 10, "bold"), command=relatorios).place(x=170, y=10, width=150, height=120)
    tk.Button(root, text="Usuarios", fg="#00BFFF", bg="#222222", border=0, font=("consolas", 10, "bold"), command=usuarios).place(x=170, y=570, width=150, height=110)
    for i in range(8, 8):
        tk.Button(root, text=f"Botão {i}", command=lambda i=i: abrir_nova_janela(f"Botão {i}"))

    # Rodar a aplicação
    root.mainloop()
# Chamar a função para criar a janela principal
criar_janela_principal()
