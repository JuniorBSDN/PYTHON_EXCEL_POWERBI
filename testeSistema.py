import tkinter as tk
from tkinter import messagebox, ttk, END
import datetime
from tkcalendar import Calendar
import openpyxl
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
import os
import pandas as pd
import pywhatkit as kit

# MENU PRINCIPAL
def criar_janela_principal():
    root = tk.Tk()
    root.title("BACK SISTEMAS ")
    root.geometry("1200x690")
    root.configure(bg="black")

    # Função para cadastrar despesas
    def cadastrar_despesas():
        root.withdraw()
        nova_janela = tk.Toplevel(root)
        nova_janela.title("Cadastrar Despesas")
        nova_janela.geometry("1400x750")
        nova_janela.configure(bg="#222222")

        def abrir_powerbi():
            caminho_powerbi = "despesas.pbix"
            try:
                os.startfile(caminho_powerbi)
            except FileNotFoundError:
                print("Arquivo Power BI não encontrado.")
            except OSError:
                print("Erro ao abrir o arquivo.")

        def abrir_tabela_despesas():
            tabela = "despesas.xlsx"
            os.startfile(tabela)

        def limpar_campos():
            entrada_descricao.delete(0, tk.END)
            entrada_quantidade.delete(0, tk.END)
            entrada_valor.delete(0, tk.END)
            entrada_parcelas.delete(0, tk.END)
            entry_dataVencimento.delete(0, tk.END)
            entry_valor_total.delete(0, tk.END)

        def gerar_txt_despesas():
            data = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            descriao = entrada_descricao.get()
            quantidade = entrada_quantidade.get()
            valor_uni = entrada_valor.get()
            parcelas = entrada_parcelas.get()
            data_vencimento = entry_dataVencimento.get()
            valor_total = float(quantidade) * float(valor_uni)
            valor_parcela = float(valor_total) / float(parcelas)

            with open('despesas.txt', 'a') as arquivo:
                arquivo.write('\n')
                arquivo.write("BACK INFORMÁTICA\n")
                arquivo.write("soluções tecnologicas\n")
                arquivo.write("fone: 91 983252639\n")
                arquivo.write("------------------------------------\n")
                arquivo.write(f"Data da transação: {data}\n")
                arquivo.write(f"Descrição: {descriao}\n")
                arquivo.write(f"Quantidade: {quantidade}\n")
                arquivo.write(f"Valor unitário: {valor_uni}\n")
                arquivo.write("------------------------------------\n")
                arquivo.write(f"Parcelas: {parcelas}\n")
                arquivo.write(f"Data de vencimento: {data_vencimento}\n")
                arquivo.write(f"Valor total: {valor_total}\n")
                arquivo.write(f"Valor parcela: {valor_parcela}\n")
                arquivo.write("\n")
                arquivo.write("**************************************")

            messagebox.showinfo("Sucesso", "Arquivo gerado com sucesso!")

        def salvar_dados_despesa():
            data_entrada = datetime.datetime.now().strftime("%Y-%m-%d")
            descricao = entrada_descricao.get()
            quantidade = int(entrada_quantidade.get())
            valorUN = float(entrada_valor.get())
            parcelas = int(entrada_parcelas.get())
            data_vencimento = entry_dataVencimento.get()
            valor_total = float(valorUN) * float(quantidade)
            valor_parcela = float(valor_total) / float(parcelas)
            try:
                # Verifica se o arquivo já existe
                workbook = load_workbook('despesas.xlsx')
                sheet = workbook.active

            except FileNotFoundError:
                # Cria um novo arquivo se não existir
                workbook = Workbook()
                sheet = workbook.active
                sheet.append(['data_entrada', 'descrição', 'quantidade', 'valorUN', 'parcelas', 'Data_vencimento','valor_total','valor_parcela'])

            # Adiciona uma nova linha com os INSTITUICAO_ENSINO
            sheet.append([data_entrada, descricao, quantidade, valorUN, parcelas, data_vencimento,valor_total,valor_parcela])
            workbook.save('despesas.xlsx')

            mensagem = (tk.messagebox.showinfo("Beleza!", "Despesa cadastrada, clique OK para adicionar outra despesa!"))

            gerar_txt_despesas()
            limpar_campos()

        def buscar_data():
            data_selecionada = cal.get_date()
            entry_dataVencimento.delete(0, tk.END)
            entry_dataVencimento.insert(0, data_selecionada)

        def buscar_hora():
            hora = datetime.datetime.now().strftime("%H:%M:%S")
            entry_dataEntrada.delete(0, tk.END)
            entry_dataEntrada.insert(0, hora)

        def carregar_tabela():
            try:
                caminho_arquivo = "despesas.xlsx"

                if caminho_arquivo:
                    df = pd.read_excel(caminho_arquivo)
                    for index, row in df.iterrows():
                        tabela.insert("", tk.END, values=list(row))

                    # Atualiza a variável global com o DataFrame
                    global dados_tabela
                    dados_tabela = df

            except Exception as e:
                messagebox.showwarning("erro inesperado!")

        def selecionar_item(event):
            try:
                item_selecionado = tabela.selection()[0]  # Obtem o item selecionado
                valores = tabela.item(item_selecionado)['values']  # Obtem os valores das colunas
                entry_dataEntrada.delete(0, tk.END)  # Limpa os campos de entrada
                entry_dataEntrada.insert(0, valores[0])  # Assumindo que a coluna "data_entrada" seja a primeira
                entrada_descricao.delete(0, tk.END)
                entrada_descricao.insert(0, valores[1])  # Assumindo que a coluna "descrição" seja a segunda
                entrada_quantidade.delete(0, tk.END)
                entrada_quantidade.insert(0, valores[2])  # Assumindo que a coluna "quantidade" seja a terceira
                entrada_valor.delete(0, tk.END)
                entrada_valor.insert(0, valores[3])  # Assumindo que a coluna "valorUN" seja a quarta
                entrada_parcelas.delete(0, tk.END)
                entrada_parcelas.insert(0, valores[4])  # Assumindo que a coluna "parcelas" seja a quinta
                entry_dataVencimento.delete(0, tk.END)
                entry_dataVencimento.insert(0, valores[5])  # Assumindo que a coluna "Data_vencimento" seja a sexta

                entry_valor_total.delete(0, tk.END)
                entry_valor_total.insert(0, valores[6])  # Assumindo que a coluna "valor_total" seja a setima
            except IndexError:
                # Lida com o caso em que nenhum item está selecionado
                messagebox.showwarning("Aviso", "Nenhum item selecionado.")

        def atualizar_despesa():
            try:
                item_selecionado = tabela.selection()[0]  # Obtem o item selecionado
                nova_data = datetime.datetime.now().strftime("%d/%m/%Y")  # obtem dados atuais da hora
                nova_despesa = entrada_descricao.get()
                nova_quantidade = entrada_quantidade.get()
                novo_valor_uni = entrada_valor.get()
                novo_parcelas = entrada_parcelas.get()
                novo_data_vencimento = entry_dataVencimento.get()
                novo_valor_total = float(novo_valor_uni) * float(nova_quantidade)

                # Atualiza os dados na tabela
                tabela.item(item_selecionado,
                            values=[nova_data, nova_despesa, novo_valor_uni, nova_quantidade, novo_parcelas,
                                    novo_data_vencimento, novo_valor_total])

                indice = tabela.index(item_selecionado)
                dados_tabela.loc[indice, "data_entrada"] = nova_data
                dados_tabela.loc[indice, "descrição"] = nova_despesa
                dados_tabela.loc[indice, "quantidade"] = nova_quantidade
                dados_tabela.loc[indice, "valorUN"] = novo_valor_uni
                dados_tabela.loc[indice, "parcelas"] = novo_parcelas
                dados_tabela.loc[indice, "Data_vencimento"] = novo_data_vencimento
                dados_tabela.loc[indice, "valor_total"] = novo_valor_total

                # Salva as alterações no arquivo Excel
                dados_tabela.to_excel("despesas.xlsx",
                                      index=False)  # index=False para nao aparece novos campos de indices

            except IndexError:
                messagebox.showwarning("Aviso", "erro na atualização.")
            limpar_campos()

        def excluir_item():
            try:
                item_selecionado = tabela.selection()[0]  # Obtem o item selecionado
                tabela.delete(item_selecionado)

                # Remove o item do DataFrame
                indice = tabela.index(item_selecionado)# obtem o indice
                dados_tabela = item_selecionado.drop(indice)

                # Salva as alterações no arquivo Excel
                dados_tabela.to_excel("despesas.xlsx",
                                      index=False)  # index=False para nao aparece novos campos de indices
            except IndexError:
                messagebox.showwarning("Aviso", "erro na exclusão.")
            limpar_campos()

        #_________________________BANNER DA JANELA PRINCIPAL_______________________________
        botao_voltar = tk.Button(nova_janela, text=" << ", font=("consolas", 12), bg="#222222", bd=0, fg="#00BFFF",command=lambda: voltar_para_janela_principal(nova_janela))
        botao_voltar.place(x=20, y=20, width=20, height=20)

        banner = tk.Label(nova_janela, text="Cadastro de Despesas", bg="#222222", fg="#00BFFF", font=("System", 25, "bold"))
        banner.place(x=55, y=10)
        # MOSTRAR A DATA E HORA ATUAL
        entry_dataEntrada = tk.Entry(nova_janela, bd=0, bg="#222222", fg="white", font=("system", 12, "bold"))
        entry_dataEntrada.place(x=60, y=55, width=150, height=40)

        #_______________________PRIMEIRA LINHA DO FORMULARIO____________________________________
        # formulario despesas
        label_quantidade = tk.Label(nova_janela, text="Qnt. ", bg="#222222", fg="white", font=("consolas", 12, "bold"))
        label_quantidade.place(x=20, y=105)
        entrada_quantidade = tk.Entry(nova_janela, bg="black", bd=0, fg="white", font=("consolas", 12, "bold"))
        entrada_quantidade.place(x=20, y=130, width=50, height=40)

        label_descricao = tk.Label(nova_janela, text="Descr. ", bg="#222222", fg="white", font=("consolas", 12, "bold"))
        label_descricao.place(x=95, y=105)
        entrada_descricao = tk.Entry(nova_janela, bg="black", bd=0, fg="white", font=("consolas", 12, "bold"))
        entrada_descricao.place(x=95, y=130, width=500, height=40)

        # _______________________SEGUNDA LINHA DO FORMULARIO____________________________________
        label_valor = tk.Label(nova_janela, text="Valor", bg="#222222", fg="white", font=("consolas", 12, "bold"))
        label_valor.place(x=20, y=190)
        entrada_valor = tk.Entry(nova_janela, bg="black", bd=0, fg="white", font=("consolas", 12, "bold"))
        entrada_valor.place(x=20, y=215, width=100, height=40)

        label_parcelas = tk.Label(nova_janela, text="Parcelas", bg="#222222", fg="white",font=("consolas", 12, "bold"))
        label_parcelas.place(x=140, y=190)
        entrada_parcelas = tk.Entry(nova_janela, bg="black", bd=0, fg="white", font=("consolas", 12, "bold"))
        entrada_parcelas.place(x=140, y=215, width=150, height=40)

        # entrada da data de vencimento
        label_data_vencimento = tk.Label(nova_janela, text="data_vencimento", bg="#222222", fg="white", font=("consolas" , 12, "bold"))
        label_data_vencimento.place(x=310, y=190)

        entry_dataVencimento = tk.Entry(nova_janela, bg="black", bd=0, fg="white", font=("consolas", 12, "bold"))
        entry_dataVencimento.place(x=310, y=215, width=160, height=40)

        total_com_parcelas = tk.Label(nova_janela, text="Total c/ parcelas", font=("consolas", 12), bg="#222222", fg="white");total_com_parcelas.place(x=480,y=190)
        entry_valor_total = tk.Entry(nova_janela, bg="black", bd=0, fg="white", font=("consolas", 12, "bold"))
        entry_valor_total.place(x=480, y=215, width=160, height=40)

        # Botão para salvar a despesa
        botao_salvar = tk.Button(nova_janela, text="Salvar Despesa", command=salvar_dados_despesa, bg="white", fg="#222222", font=("consolas", 12, "bold"))
        botao_salvar.place(x=20, y=290, width=200, height=50)

        botao_atualizar = tk.Button(nova_janela, text="Atualizar", command= atualizar_despesa, bg="white", fg="#222222", font=("consolas", 12, "bold"))
        botao_atualizar.place(x=230, y=290, width=200, height=50)

        botao_excluir = tk.Button(nova_janela, text="Excluir", command= excluir_item, bg="white", fg="#222222", font=("consolas", 12, "bold"))
        botao_excluir.place(x=440, y=290, width=200, height=50)
        # calendário com botão para selecionar a data0
        label_info = tk.Label(nova_janela, text="Selecione a Data de Vencimento", bg="#222222", fg="white", font=("consolas", 12, "bold"))
        label_info.place(x=710, y=70)
        cal = Calendar(nova_janela, selectmode='day', year=2023, month=11, day=22)
        cal.place(x=730, y=100)
        select_data_vencimento = tk.Button(nova_janela, text="Selecionar", command=buscar_data, bg="#444444", bd=0,fg="#00BFFF", font=("consolas", 12, "bold"))
        select_data_vencimento.place(x=790, y=310, width=150, height=50)

        # FRAME PARA ALOCAR A TABELA
        frame = tk.Frame(nova_janela, bg="#222222")
        frame.place(x=20, y=370, width=1300, height=280)
        tabela = ttk.Treeview(frame, columns=("data_entrada", "descrição", "quantidade", "valorUN", "parcelas", "Data_vencimento","valor_total","valor_parcela"), show="headings")
        tabela.heading("data_entrada", text="DataEntrada")
        tabela.heading("descrição", text=" Descr.  ")
        tabela.heading("quantidade", text=" Qtd. ")
        tabela.heading("valorUN", text="Valor_UN")
        tabela.heading("parcelas", text="Parcelas")
        tabela.heading("Data_vencimento", text="Data_Venc")
        tabela.heading("valor_total", text="Valor_Total")
        tabela.heading("valor_parcela", text="Valor_Parcela")
        tabela.pack(fill="both", expand=True)
        tabela.bind("<Double-1>", selecionar_item)

        botao_analise_gastos = tk.Button(nova_janela, text="Analisar Gastos", bg="#444444", bd=0, fg="#00BFFF", font=("consolas", 12, "bold"), command=abrir_powerbi)
        botao_analise_gastos.place(x=20, y=660, width=400, height=50)

        botao_view = tk.Button(nova_janela, text="Visualizar Tabela", bg="#444444", bd=0, fg="#00BFFF",font=("consolas", 12, "bold"), command=abrir_tabela_despesas)
        botao_view.place(x=430, y=660, width=400, height=50)

        buscar_hora()
        carregar_tabela()

    # Função cadastrar vendas
    def cadastrar_vendas():
        root.withdraw()  # Oculta a janela principal
        nova_janela = tk.Toplevel(root)
        nova_janela.title("Cadastrar Vendas")
        nova_janela.geometry("1250x750")
        nova_janela.configure(bg="#222222")

        def carregar_tabela_vendas():
            try:
                caminho_arquivo = "vendas.xlsx"

                if caminho_arquivo:
                    df = pd.read_excel(caminho_arquivo)
                    for index, row in df.iterrows():
                        tabela.insert("", tk.END, values=list(row))

                    # Atualiza a variável global com o DataFrame
                    global dados_tabela
                    dados_tabela = df

            except Exception as e:
                print(f"Erro ao carregar a tabela: {e}")

        def analise_vendas():
            caminho_powerbi = "vendas.pbix"
            try:
                os.startfile(caminho_powerbi)
            except FileNotFoundError:
                print("Arquivo Power BI não encontrado.")
            except OSError:
                print("Erro ao abrir o arquivo.")

        def abrir_tabela_vendas():
            tabela = "vendas.xlsx"
            os.startfile(tabela)

        def limpar_campos():
            entrada_item.delete(0, tk.END)
            entrada_descricao.delete(0, tk.END)
            entrada_nova_quantidade.delete(0, tk.END)
            entrada_valor_uni.delete(0, tk.END)
            entrada_valor_pago.delete(0, tk.END)
            entrada_troco.delete(0,tk.END)

        def selecionar_item(event):
            try:
                item_selecionado = tabela.selection()[0]
                valores = tabela.item(item_selecionado)['values']
                entry_dataEntrada_venda.delete(0, tk.END)
                entry_dataEntrada_venda.insert(0, valores[0])
                entrada_item.delete(0, tk.END)
                entrada_item.insert(0, valores[1])
                entrada_descricao.delete(0, tk.END)
                entrada_descricao.insert(0, valores[2])
                entrada_nova_quantidade.delete(0, tk.END)
                entrada_nova_quantidade.insert(0, valores[3])
                entrada_valor_uni.delete(0, tk.END)
                entrada_valor_uni.insert(0, valores[4])
                entrada_valor_pago.delete(0, tk.END)
                entrada_valor_pago.insert(0, valores[5])

                entrada_troco.delete(0, tk.END)
                entrada_troco.insert(0, valores[6])


            except:
                print("erro ao selecionar registro")

        def alterar_venda():
            try:
                item_selecionado = tabela.selection()[0]
                nova_data = datetime.datetime.now().strftime("%d/%m/%y")
                novo_item = entrada_item.get()
                nova_descri = entrada_descricao.get()
                nova_quant = entrada_nova_quantidade.get()
                novo_valor_uni = entrada_valor_uni.get()
                novo_valor_pago = entrada_valor_pago.get()
                novo_valor_total = float(novo_valor_uni)*float(nova_quant)
                novo_troco = float(novo_valor_pago) - float(novo_valor_total)

                tabela.item(item_selecionado, values=[nova_data, novo_item, nova_descri, nova_quant, novo_valor_uni, novo_valor_pago, novo_valor_total, novo_troco ])

                indice = tabela.index(item_selecionado)
                dados_tabela.loc[indice, "Data da transação"] = nova_data
                dados_tabela.loc[indice, "Item"] = novo_item
                dados_tabela.loc[indice, "Descrição"] = nova_descri
                dados_tabela.loc[indice, "Quantidade"] = nova_quant
                dados_tabela.loc[indice, "Valor unitário"] = novo_valor_uni
                dados_tabela.loc[indice, "Valor pago"] = novo_valor_pago
                dados_tabela.loc[indice, "Total"] = novo_valor_total
                dados_tabela.loc[indice, "Troco"] = novo_troco

                dados_tabela.to_excel("vendas.xlsx", index=False)
            except:
                messagebox.showwarning("Erro na alteração")
            limpar_campos()

        def excluir_venda():
            # Implementar a lógica para excluir um registro selecionado
            # e atualizar o arquivo Excel
            pass

        def gerar_txt():
            data = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            item = entrada_item.get()
            descricao = entrada_descricao.get()
            quantidade = entrada_nova_quantidade.get()
            valor_uni = entrada_valor_uni.get()
            valor_pago = entrada_valor_pago.get()
            valor_total = float(valor_uni) * float(quantidade)
            troco = float(valor_pago) - float(valor_total)

            with open("vendas.txt", "a") as f:
                f.write("\n")
                f.write(f"NOME DA EMPRESA: XxX\n")
                f.write(f"soluções tecnologicas\n")
                f.write(f"fone: 91 983252639\n")
                f.write(f"------------------------------------\n")
                f.write(f"Data da transação: {data}\n")
                f.write(f"Item: {item}\n")
                f.write(f"Descrição: {descricao}\n")
                f.write(f"Quantidade: {quantidade}\n")
                f.write(f"Valor unitário: {valor_uni}\n")
                f.write(f"------------------------------------\n")
                f.write(f"Valor pago: {valor_pago}\n")
                f.write(f"Total: {valor_total}\n")
                f.write(f"Troco: {troco}\n")
                f.write("\n")
                f.write(f"Muito obrigado por usar nosso TESTE SISTEMA\n")
                f.write("\n")
                f.write(f"* * * * * * * * * * * * * * * * * * * *")
            messagebox.showinfo("Sucesso", "Arquivo gerado com sucesso!")

        def salvar_vendas():
            nova_data = entry_dataEntrada_venda.get()
            novo_item = entrada_item.get()
            nova_descricao = entrada_descricao.get()
            nova_quantidade = entrada_nova_quantidade.get()
            novo_valor_uni = entrada_valor_uni.get()
            novo_valor_pago = entrada_valor_pago.get()
            novo_valor_total = float(novo_valor_uni) * float(nova_quantidade)
            troco = float(novo_valor_pago) - float(novo_valor_total)

            try:
                # Verifica se o arquivo já existe
                workbook = load_workbook('vendas.xlsx')
                sheet = workbook.active
            except FileNotFoundError:
                # Cria um novo arquivo se não existir
                workbook = Workbook()
                sheet = workbook.active
                sheet.append(
                    ['Data da transação', 'Item', 'Descrição', 'Quantidade', 'Valor unitário', 'Valor pago', 'Total',
                     'Troco'])
            # Adiciona uma nova linha
            sheet.append([nova_data, novo_item, nova_descricao, nova_quantidade, novo_valor_uni, novo_valor_pago,novo_valor_total, troco])
            workbook.save('vendas.xlsx')

            # Exibir a mensagem de confirmação
            mensagem = (tk.messagebox.showinfo("OK!", "Despesa cadastrada, clique OK para adicionar outra despesa."))
            gerar_txt()
            limpar_campos()

        def buscar_hora():
            hora = datetime.datetime.now().strftime("%H:%M:%S")
            entry_dataEntrada_venda.delete(0, tk.END)
            entry_dataEntrada_venda.insert(0, hora)

        # _________________________BANNER DA JANELA PRINCIPAL_______________________________
        botao_voltar = tk.Button(nova_janela, text=" << ", font=("consolas", 12), bg="#222222", bd=0, fg="#00BFFF",command=lambda: voltar_para_janela_principal(nova_janela))
        botao_voltar.place(x=20, y=20, width=20, height=20)

        banner = tk.Label(nova_janela, text="Cadastro de Vendas", bg="#222222", fg="#00BFFF",font=("System", 25, "bold"))
        banner.place(x=55, y=10)

        # MOSTRAR A DATA E HORA ATUAL
        entry_dataEntrada_venda = tk.Entry(nova_janela, bg="#222222", bd=0, fg="white", font=("consolas", 12, "bold"))
        entry_dataEntrada_venda.place(x=80, y=55, width=100, height=20)

        #____________________________PRIMEIRA LINHA DO FORMULARIO____________________________
        label_item = tk.Label(nova_janela, text="Item", font=("consolas", 12), bg="#222222", fg="white");label_item.place(x=20, y=100, width=50, height=40)
        entrada_item = tk.Entry(nova_janela, font=("consolas", 12), bg="black", bd=0, fg="white");entrada_item.place(x=20, y=130, width=150, height=40)

        label_descri = tk.Label(nova_janela, text="Descrição", font=("consolas", 12), bg="#222222", fg="white"); label_descri.place(x=190, y=100, width=120, height=40)
        entrada_descricao = tk.Entry(nova_janela, font=("consolas", 12), bg="black", bd=0, fg="white"); entrada_descricao.place(x=190, y=130, width=400, height=40)

        #_____________________________SEGUNDA LINHA DO FORMULARIO____________________________
        labe_quant = tk.Label(nova_janela, text="Qtd.: ", font=("consolas", 12), bg="#222222", fg="white");labe_quant.place(x=20, y=180, width=80, height=40)
        entrada_nova_quantidade = tk.Entry(nova_janela, font=("consolas", 12), bg="black", bd=0, fg="white"); entrada_nova_quantidade.place(x=20, y=210, width=80, height=40)

        label_valor_uni = tk.Label(nova_janela, text="ValorUnitário:", font=("consolas", 12), bg="#222222", fg="white");label_valor_uni.place(x=130, y=180, width=160, height=40)
        entrada_valor_uni = tk.Entry(nova_janela, font=("consolas", 12), bg="black", bd=0, fg="white");entrada_valor_uni.place(x=130, y=210, width=160, height=40)

        label_valor_pago = tk.Label(nova_janela, text="valorPago:", font=("consolas", 12), bg="#222222", fg="white");label_valor_pago.place(x=310, y=180, width=150, height=40)
        entrada_valor_pago = tk.Entry(nova_janela, font=("consolas", 12), bg="black", bd=0, fg="white");entrada_valor_pago.place(x=320, y=210, width=150, height=40)

        label_troco = tk.Label(nova_janela, text="troco", font=("consolas", 12), bg="#222222", fg="white");label_troco.place(x=500,y=180)
        entrada_troco = tk.Entry(nova_janela, font=("consolas, 12 "), bg="black", bd=0, fg="white"); entrada_troco.place(x=500, y=210, width =150, height=40)

        #BOTÃO PARA SALVAR A VENDA
        botao_inserir = tk.Button(nova_janela, text="Salvar venda",  font=("consolas", 12), bg="white", bd=0, fg="#222222", command=salvar_vendas)
        botao_inserir.place(x=20, y=280, width=200, height=40)

        botao_altetrar = tk.Button(nova_janela, text="Alterar venda",  font=("consolas", 12), bg="white", bd=0, fg="#222222", command=alterar_venda)
        botao_altetrar.place(x=240, y=280, width=200, height=40)

        botao_excluir = tk.Button(nova_janela, text="Excluir venda",  font=("consolas", 12), bg="white", bd=0, fg="#222222", command=excluir_venda)
        botao_excluir.place(x=460, y=280, width=200, height=40)

        botao_analise = tk.Button(nova_janela, text="Análise de Vendas", font=("consolas", 12), bg="#444444", bd=0, fg="#00BFFF",command=analise_vendas)
        botao_analise.place(x=20, y=620, width=400, height=50)

        botao_view = tk.Button(nova_janela, text="Visualizar Tabela", font=("consolas", 12), bg="#444444", bd=0, fg="#00BFFF",command=abrir_tabela_vendas)
        botao_view.place(x=430, y=620, width=400, height=50)

        frame = tk.Frame(nova_janela, bg="#222222")
        frame.place(x=20, y=350, width=1220, height=250)

        tabela = ttk.Treeview(frame, columns=('Data da transação', 'Item', 'Descrição', 'Quantidade', 'Valor unitário', 'Valor pago', 'Total','Troco'),show="headings")
        tabela.heading("Data da transação", text="Data da transação")
        tabela.heading("Item", text="Item")
        tabela.heading("Descrição", text="Descrição")
        tabela.heading("Quantidade", text="Quantidade")
        tabela.heading("Valor unitário", text="Valor unitário")
        tabela.heading("Valor pago", text="Valor pago")
        tabela.heading("Total", text="Total")
        tabela.heading("Troco", text="Troco")

        tabela.column("Data da transação")
        tabela.column("Item")
        tabela.column("Descrição")
        tabela.column("Quantidade")
        tabela.column("Valor unitário")
        tabela.column("Valor pago")
        tabela.column("Total")
        tabela.column("Troco")
        tabela.pack(fill="both", expand=True)

        tabela.bind("<Double-1>", selecionar_item)

        buscar_hora()
        carregar_tabela_vendas()

    #função para adicionar o Estoque
    def add_estoque():
        root.withdraw()  # Oculta a janela principal
        nova_janela = tk.Toplevel(root)
        nova_janela.title("Cadastrar Estoque")
        nova_janela.geometry("1250x750")
        nova_janela.configure(bg="#222222")

        def carregar_tabela():
            try:
                caminho_arquivo = "estoque.xlsx"

                if caminho_arquivo:
                    df = pd.read_excel(caminho_arquivo)
                    for index, row in df.iterrows():
                        tabela.insert("", tk.END, values=list(row))

                    # Atualiza a variável global com o DataFrame
                    global dados_tabela
                    dados_tabela = df

            except Exception as e:
                messagebox.showwarning("erro inesperado!")

        def selecionar_item(event):
            try:
                item_selecionado = tabela.selection()[0]
                item = tabela.item(item_selecionado)
                codigo_barras_entry.delete(0, tk.END)
                descricao_entry.delete(0, tk.END)
                quantidade_entry.delete(0, tk.END)
                valor_inicial_entry.delete(0, tk.END)
                valor_final_entry.delete(0, tk.END)
                entry_lucro.delete(0, tk.END)
                entry_lucro.insert(0, item["values"][5])
                entry_lucro_total.delete(0, tk.END)
                entry_lucro_total.insert(0, item["values"][6])

                codigo_barras_entry.insert(0, item["values"][0])
                descricao_entry.insert(0, item["values"][1])
                quantidade_entry.insert(0, item["values"][2])
                valor_inicial_entry.insert(0, item["values"][3])
                valor_final_entry.insert(0, item["values"][4])
                entry_lucro.insert(0, item["values"][5])
                entry_lucro_total.insert(0, item["values"][6])

            except IndexError:
                messagebox.showwarning("erro inesperado!")

        def alterar_estoque():
            try:
                item_selecionado = tabela.selection()[0]
                codigo_barras = codigo_barras_entry.get()
                descricao = descricao_entry.get()
                quantidade = quantidade_entry.get()
                valor_inicial = valor_inicial_entry.get()
                valor_final = valor_final_entry.get()
                lucro = float(valor_final) - float(valor_inicial)
                lucroTotal = float(valor_final) * float(quantidade)
                tabela.item(item_selecionado, values=(codigo_barras, descricao, quantidade, valor_inicial, valor_final, lucro, lucroTotal))

                indice = tabela.index(item_selecionado)
                dados_tabela.loc[indice, "Código de Barras"] = codigo_barras
                dados_tabela.loc[indice, "Descrição"] = descricao
                dados_tabela.loc[indice, "Quantidade"] = quantidade
                dados_tabela.loc[indice, "Valor Inicial"] = valor_inicial
                dados_tabela.loc[indice, "Valor Final"] = valor_final
                dados_tabela.loc[indice, "Lucro"] = lucro
                dados_tabela.loc[indice, "Lucro Total"] = lucroTotal

                dados_tabela.to_excel("estoque.xlsx", index=False)

            except IndexError:
                messagebox.showwarning("erro inesperado!")

        def excluir_estoque():
            pass

        def analise_estoque():
            caminho_powerbi = "estoque.pbix"
            try:
                os.startfile(caminho_powerbi)
            except FileNotFoundError:
                print("Arquivo Power BI não encontrado.")
            except OSError:
                print("Erro ao abrir o arquivo.")

        def abrir_tabela_estoque():
            tabela = "estoque.xlsx"
            os.startfile(tabela)
            print("Tabela aberta com sucesso!")

        def limpar_campos():
            codigo_barras_entry.delete(0, tk.END)
            descricao_entry.delete(0, tk.END)
            quantidade_entry.delete(0, tk.END)
            valor_inicial_entry.delete(0, tk.END)
            valor_final_entry.delete(0, tk.END)
            entry_lucro.delete(0, tk.END)
            entry_lucro_total.delete(0, tk.END)
            print("Campos limpos com sucesso!")

        def salvar_estoque():
            codigo_barras = codigo_barras_entry.get()
            descricao = descricao_entry.get()
            quantidade = quantidade_entry.get()
            valor_inicial = valor_inicial_entry.get()
            valor_final = valor_final_entry.get()
            lucro = float('valor_final') - float('valor_inicial')
            lucroTotal = float("valor_final") * float("quantidade")

            try:
                workbook = openpyxl.load_workbook('estoque.xlsx')
                codigo_barras = codigo_barras_entry.get()
                descricao = descricao_entry.get()
                quantidade = quantidade_entry.get()
                valor_inicial = valor_inicial_entry.get()
                valor_final = valor_final_entry.get()
                lucro = float(valor_final) - float(valor_inicial)
                lucroTotal = float(valor_final) * float(quantidade)
                sheet = workbook.active

            except FileNotFoundError:
                workbook = openpyxl.Workbook()
                sheet = workbook.active
                sheet.append(['Código de Barras', 'Descrição', 'Quantidade', 'Valor Inicial', 'Valor Final', 'Lucro',
                              'lucroTotal'])

            # Adicionar uma nova linha da tabela  produto
            sheet.append([codigo_barras, descricao, quantidade, valor_inicial, valor_final, lucro, lucroTotal])

            # Salvar o arquivo
            workbook.save('estoque.xlsx')
            limpar_campos()
            print("Estoque salvo com sucesso!")

        # _________________________BANNER DA JANELA PRINCIPAL_______________________________
        botao_voltar = tk.Button(nova_janela, text=" << ", font=("consolas", 12), bg="#222222", bd=0, fg="#00BFFF", command=lambda: voltar_para_janela_principal(nova_janela))
        botao_voltar.place(x=20, y=20, width=20, height=20)

        banner = tk.Label(nova_janela, text="Cadastro de Estoque", bg="#222222", fg="#00BFFF", font=("System", 25, "bold")); banner.place(x=55, y=10)

        label_codigo = tk.Label(nova_janela, text="Cod.", fg="white", bg="#222222", font=("consolas", 12, "bold")); label_codigo.place(x=20, y=110)
        codigo_barras_entry = tk.Entry(nova_janela, font=("consolas", 12), bg="black", bd=0, fg="white"); codigo_barras_entry.place(x=20, y=135, width=60, height=40)

        label_quantidade = tk.Label(nova_janela, text="Qtd.", fg="white", bg="#222222", font=("consolas", 12, "bold")); label_quantidade.place(x=100, y=110)
        quantidade_entry = tk.Entry(nova_janela, font=("consolas", 12), bg="black", bd=0, fg="white"); quantidade_entry.place(x=100, y=135, width=60, height=40)

        label_descricao = tk.Label(nova_janela, text="Desc.", fg="white", bg="#222222", font=("consolas", 12, "bold")); label_descricao.place(x=185, y=110)
        descricao_entry = tk.Entry(nova_janela, font=("consolas", 12), bg="black", bd=0, fg="white"); descricao_entry.place(x=185, y=135, width=300, height=40)

        label_valor_uni = tk.Label(nova_janela, text="valorInicial", fg="white", bg="#222222", font=("consolas", 12, "bold")); label_valor_uni.place(x=20, y=200)
        valor_inicial_entry = tk.Entry(nova_janela, font=("consolas", 12), bg="black", bd=0, fg="white"); valor_inicial_entry.place(x=20, y=225, width=150, height=40)

        label_valor_final = tk.Label(nova_janela, text="Valor_final", fg="white", bg="#222222", font=("consolas", 12, "bold")); label_valor_final.place(x=190, y=200)
        valor_final_entry = tk.Entry(nova_janela, font=("consolas", 12), bg="black", bd=0, fg="white"); valor_final_entry.place(x=190, y=225, width=150, height=40)

        label_lucro = tk.Label(nova_janela, text="Lucro", fg="white", bg="#222222", font=("consolas", 12, "bold")); label_lucro.place(x=360, y=200)
        entry_lucro = tk.Entry(nova_janela, font=("consolas", 12), bg="black", bd=0, fg="white");entry_lucro.place(x=360, y=225, width=150, height=40)

        label_lucro_total = tk.Label(nova_janela, text="Lucro Total", fg="white", bg="#222222", font=("consolas", 12, "bold"));label_lucro_total.place(x=530, y=200)
        entry_lucro_total = tk.Entry(nova_janela, font=("consolas", 12), bg="black", bd=0, fg="white");entry_lucro_total.place(x=530, y=225, width=150, height=40)

        botao_inserir = tk.Button(nova_janela, text=" Adicionar ", font=("consolas", 12), bg="white", bd=0, fg="#222222", command=salvar_estoque);botao_inserir.place(x=20, y=300, width=400, height=40)

        botao_atualizar = tk.Button(nova_janela, text=" Atualizar ", font=("consolas", 12), bg="white", bd=0, fg="#222222", command=alterar_estoque);botao_atualizar.place(x=430, y=300, width=400, height=40)

        botao_excluir = tk.Button(nova_janela, text=" Excluir ", font=("consolas", 12), bg="white", bd=0, fg="#222222", command=excluir_estoque);botao_excluir.place(x=550, y=380, width=400, height=40)

        frame = tk.Frame(nova_janela, bg="#222222")
        frame.place(x=20, y=380, width=1220, height=250)
        tabela = ttk.Treeview(frame,columns=('codigo_barras', 'descricao', 'quantidade', 'valor_inicial', 'valor_final','lucro', 'lucroTotal'), show="headings")
        tabela.heading('codigo_barras', text='Código de Barras')
        tabela.heading('descricao', text='Descrição')
        tabela.heading('quantidade', text='Quantidade')
        tabela.heading('valor_inicial', text='Valor Inicial')
        tabela.heading('valor_final', text='Valor Final')
        tabela.heading('lucro', text='Lucro')
        tabela.heading('lucroTotal', text='Lucro Total')
        tabela.pack(fill="both", expand=True)
        tabela.bind("<Double-1>", selecionar_item)

        botao_analise = tk.Button(nova_janela, text="Análisa Estoque", font=("consolas", 12), bg="#444444", bd=0, fg="#00BFFF",command=analise_estoque)
        botao_analise.place(x=20, y=650, width=400, height=50)

        botao_view = tk.Button(nova_janela, text="Visualizar Tabela", font=("consolas", 12), bg="#444444", bd=0, fg="#00BFFF",command=abrir_tabela_estoque)
        botao_view.place(x=430, y=650, width=400, height=50)

        carregar_tabela()

    # _________________________JANELA RELATORIOS_______________________________
    def relatorios():
        root.withdraw()  # Oculta a janela principal
        nova_janela = tk.Toplevel(root)
        nova_janela.title("Relatórios")
        nova_janela.geometry("1200x700")
        nova_janela.configure(bg="black")



        # _________________________BANNER DA JANELA PRINCIPAL_______________________________
        botao_voltar = tk.Button(nova_janela, text=" << ", font=("consolas", 12, "bold"), fg="#00BFFF", bg="black", bd=0, command=lambda: voltar_para_janela_principal(nova_janela))
        botao_voltar.place(x=20, y=20, width=20, height=20)

        banner = tk.Label(nova_janela, text="Analise e relatorios", bg="black", fg="#00BFFF", font=("System", 25, "bold"))
        banner.place(x=55, y=10)

        label_data_entrada = tk.Label(nova_janela, text=datetime.datetime.now().strftime("%d/%m/%Y   %H:%M:%S"), bg="black", fg="white", font=("System", 12, "bold"))
        label_data_entrada.place(x=60, y=60)

        frame_grafico_vendas = tk.Frame(nova_janela, bg="#222222")
        frame_grafico_vendas.place(x=400, y=100, width=250, height=480)
        label_vendas_dash = tk.Label(frame_grafico_vendas, text="Vendas", font="consolas 12 bold", bg="#222222", fg="#00bfff")
        label_vendas_dash.place(x=10, y=10)

        frame_grafico_despesas = tk.Frame(nova_janela, bg="#222222")
        frame_grafico_despesas.place(x=665, y=100, width=250, height=480)
        label_despesa_dash = tk.Label(frame_grafico_despesas, text="Despesas", font="consolas 12 bold", bg="#222222", fg="#00bfff")
        label_despesa_dash.place(x=10, y=10)

        frame_grafico_estoque = tk.Frame(nova_janela, bg="#222222")
        frame_grafico_estoque.place(x=930, y=100, width=250, height=480)
        label_estoque_dash = tk.Label(frame_grafico_estoque, text="Estoque", font="consolas 12 bold", bg="#222222", fg= "#00bfff")
        label_estoque_dash.place(x=10, y=10)

    #cadastre aqui os USUARIOS
    def usuarios():
        root.withdraw()  # Oculta a janela principal
        nova_janela = tk.Toplevel(root)
        nova_janela.title("Usuarios")
        nova_janela.configure(bg="#222222")
        nova_janela.geometry("1200x700")

        # _________________________BANNER DA JANELA PRINCIPAL_______________________________
        botao_voltar = tk.Button(nova_janela, text=" << ", font=("consolas", 12, "bold"), fg="#00BFFF", bg="#222222",bd=0, command=lambda: voltar_para_janela_principal(nova_janela))
        botao_voltar.place(x=20,y=20, width=20, height=20)

        banner = tk.Label(nova_janela, text="Usuarios", bg="#222222", fg="#00BFFF",font=("System", 25, "bold"))
        banner.place(x=55, y=10)

        # MOSTRAR A DATA E HORA ATUAL
        label_data_entrada = tk.Label(nova_janela, text=datetime.datetime.now().strftime("%d/%m/%Y %H:%M:%S"),bg="#222222", fg="white", font=("System", 10, "bold"))
        label_data_entrada.place(x=60, y=55)

        # Função para cadastrar um usuário

        def limpar_campos():
            entry_user.delete(0, tk.END)
            entry_senha.delete(0, tk.END)

        def cadastrar_usuario():
            user = entry_user.get()
            senha = entry_senha.get()

            try:
                workbook = openpyxl.load_workbook('usuarios.xlsx')
                sheet = workbook.active
                sheet.append([user, senha])
                workbook.save('usuarios.xlsx')
                messagebox.showinfo("Sucesso", "Usuário cadastrado com sucesso!")

            except Exception as e:
                workbook = openpyxl.Workbook()
                sheet = workbook.active
                sheet.append([user, senha])
                workbook.save('usuarios.xlsx')
                messagebox.showerror("Erro", f"Erro ao cadastrar usuário: {str(e)}")
            limpar_campos()

        label_user = tk.Label(nova_janela, text="Usuário:", bg="#222222", fg="white",font=("System", 12, "bold"))
        label_user.place(x=20, y=100)
        entry_user = tk.Entry(nova_janela, font=("consolas", 12), bg="black", bd=0, fg="white")
        entry_user.place(x=20, y=130)

        label_senha = tk.Label(nova_janela, text="Senha:", bg="#222222", fg="white",font=("System", 12, "bold"))
        label_senha.place(x=20, y=160)
        entry_senha = tk.Entry(nova_janela, show="*", font=("consolas", 12), bg="black", bd=0, fg="white")
        entry_senha.place(x=20, y=190)

        botão_cadastrar_usuario = tk.Button(nova_janela, text="Cadastrar",font=("consolas", 12, "bold"), fg="#222222", bg="#00BFFF", bd=0,command=cadastrar_usuario)
        botão_cadastrar_usuario.place(x=20, y=220, width=200, height=50)
        # Função para editar um usuário
        def editar_usuario():
            pass

        # ... (lógica para buscar o usuário e permitir a edição)

        # Função para remover um usuário
        def remover_usuario():
            pass

        botao_editar = tk.Button(nova_janela, text="Editar",font=("consolas", 12, "bold"), fg="#222222", bg="#00BFFF", bd=0,command=editar_usuario)
        botao_editar.place(relx=0.5, rely=0.4, anchor="center", width=200, height=50)

        botao_remover = tk.Button(nova_janela, text="Remover",font=("consolas", 12, "bold"), fg="#222222", bg="#00BFFF", bd=0,command=remover_usuario)
        botao_remover.place(relx=0.5, rely=0.5, anchor="center", width=200, height=50)

    #informações do programa
    def informa():
        root.withdraw()  # Oculta a janela principal
        nova_janela = tk.Toplevel(root)
        nova_janela.title("informações")
        nova_janela.configure(bg="#222222")
        nova_janela.geometry("1200x700")

        banner = tk.Label(nova_janela, text="BACK INFORMÁTICA\n"
                                            "Soluções tecnologicas\n"
                                            "\n"
                                            "* Esta versão 1,0 feita para cadastrar despesas,vendas e estoque em arquivos simples como \n"
                                            "o excel e para analise de dados com power bi, atravez de uma interface amigavel e tema escuro,\n"
                                            "para melhor visibilidade e conforto do usuario\n"
                                            " \n"
                                            "\n"
                                            "* desenvolvido por: AIRTON JUNIOR\n"
                                            "TECNOLOGO E ESTUDANTE DO CURSO DE "
                                            "ANANLISE E DESENVOLVIMENTO DE SOFTWARES\n"
                                            "\n"
                                            "\n"
                                            " Obrigado por usar meu TESTE SISTEMA\n"
                                    , bg="#222222", fg="red", font=("consolas", 12, "bold"))
        banner.pack()

        botao_voltar = tk.Button(nova_janela, text="Voltar", fg="white", bg="#333333", bd=0, command=lambda: voltar_para_janela_principal(nova_janela))
        botao_voltar.place(relx=0.5, rely=0.9, anchor="center", width=100, height=50)

    # Função para enviar mensagem no whatsapp
    def send_whatsapp_message():
        phone_number = "+5591983252639"
        message = "ola, pertenço ao software da serie v1.0, aguardando retorno, obrigado !"
        kit.sendwhatmsg_instantly(phone_number, message)

    #ESTAS FUNÇÓES ESTÃO EM VARIAS JANELAS DA INTERFACE, FUNDAMENTAIS.
    def abrir_nova_janela(nome_botao):
        root.withdraw()  # Oculta a janela principal
        nova_janela = tk.Toplevel(root)
        nova_janela.title(nome_botao)

        label = tk.Label(nova_janela, text=f"Você clicou no {nome_botao}")
        label.pack(pady=10)

        botao_voltar = tk.Button(nova_janela, text="Voltar", command=lambda: voltar_para_janela_principal(nova_janela))
        botao_voltar.pack(pady=10)

    # Função para voltar para a janela principal
    def voltar_para_janela_principal(janela_atual):
        janela_atual.destroy()
        root.deiconify()  # Mostra a janela principal novamente

    # função para fefchar janela principal
    def fechar():
        root.destroy()

    def somar_vendas():
        aruivo_vendas = "vendas.xlsx"
        df = pd.read_excel(aruivo_vendas)
        soma_total_vendas = df["Total"].sum()
        return soma_total_vendas

    def somar_despesas():
        aruivo_despesas = "despesas.xlsx"
        df = pd.read_excel(aruivo_despesas)
        soma_total_despesas = df["valor_total"].sum()
        return soma_total_despesas

    def somar_estoque():
        aruivo_estoque = "estoque.xlsx"
        df = pd.read_excel(aruivo_estoque)
        soma_total_estoque = df["lucro"].sum()
        return soma_total_estoque


    #INTERFACE PRINCIPAL
    frame_logo = tk.Frame(root, bg="black", bd=0)
    frame_logo.place(x=330, y=10, width=660, height=670)
    logoback = tk.PhotoImage(file="logo.png")
    label_logo = tk.Label(frame_logo, image=logoback, bg="black")
    label_logo.place(x=0, y=0, width=660, height=670)

    label_time_system = tk.Label(root, text=datetime.datetime.now().strftime("%d/%m/%Y   %H:%M:%S"), bg="black", fg="#00BFFF", font="system 12 bold")
    label_time_system.place(x=1030, y=650)

    frame_analise_venda = tk.Frame(root, bg="#222222", bd=0, )
    frame_analise_venda.place(x=1040, y=10,  width=150, height=200)
    label_resultado_vendas = tk.Label(frame_analise_venda, text="Total Venda:", fg="#00bfff", bg="#222222")
    label_resultado_vendas.place(x=10, y=10)
    label_total_vendas = tk.Label(frame_analise_venda, text="", fg="#00bfff", bg="#222222", font="consolas 12 bold" )
    label_total_vendas.place(x=10, y=40)

    label_result = tk.Label(frame_analise_venda, text=f"R$:{somar_vendas()}", fg="#00bfff", bg="#222222", font="consolas 12 bold")
    label_result.place(x=10, y=70)

    frame_analise_despesa = tk.Frame(root, bg="#222222", bd=0, )
    frame_analise_despesa.place(x=1040, y=220, width=150, height=200)
    label_resultado_despesa = tk.Label(frame_analise_despesa, text="Total despesas:", fg="#00bfff", bg="#222222")
    label_resultado_despesa.place(x=10, y=10)

    label_result1 = tk.Label(frame_analise_despesa, text=f"R$:{somar_despesas()}", fg="#00bfff", bg="#222222", font="consolas 12 bold")
    label_result1.place(x=10, y=40)

    frame_analise_estoque = tk.Frame(root, bg="#222222", bd=0, )
    frame_analise_estoque.place(x=1040, y=430, width=150, height=200)
    label_resultado_estoque = tk.Label(frame_analise_estoque, text="Total estoque:", fg="#00bfff", bg="#222222")
    label_resultado_estoque.place(x=10, y=10)

    label_resul2 = tk.Label(frame_analise_estoque, text=f"R$:{somar_estoque()}", fg="#00bfff", bg="#222222", font="consolas 12 bold")
    label_resul2.place(x=10, y=40)

    versao = tk.Label(frame_logo, text="v1.0", bg="black", fg="#00BFFF", font=("consolas", 12, "bold"))
    versao.place(x=300, y=355)

    tk.Button(root, text="Despesas", fg="#00BFFF", bg="#222222", border=0, font=("consolas", 10, "bold"), command= cadastrar_despesas).place(x=10, y=10, width=150, height=300)
    tk.Button(root, text="Vendas", fg="#00BFFF", bg="#222222", border=0, font=("consolas", 10, "bold"), command= cadastrar_vendas).place(x=10, y=380, width=150, height=300)
    tk.Button(root, text="Estoque", fg="#00BFFF", bg="#222222", border=0, font=("consolas", 10, "bold"), command= add_estoque).place(x=170, y=200, width=150, height=300)
    tk.Button(root, text=" Informações ", fg="#222222", bg="#00BFFF", border=0, font=("consolas", 10, "bold"), command= informa).place(x=170, y=140, width=150, height=50)
    tk.Button(root, text=" Sair/Fechar ", fg="#222222", bg="#00BFFF", border=0, font=("consolas", 10, "bold"), command= fechar).place(x=170, y=510, width=150, height=50)
    tk.Button(root, text="Suporte", fg="#222222", bg="#00BFFF", border=0, font=("consolas", 10, "bold"), command= send_whatsapp_message).place(x=10, y=320, width=150, height=50)
    tk.Button(root, text="Relatorios", fg="#00BFFF", bg="#222222", border=0, font=("consolas", 10, "bold"), command=relatorios).place(x=170, y=10, width=150, height=120)
    tk.Button(root, text="Usuarios", fg="#00BFFF", bg="#222222", border=0, font=("consolas", 10, "bold"), command=usuarios).place(x=170, y=570, width=150, height=110)
    for i in range(8, 8):
        tk.Button(root, text=f"Botão {i}", command=lambda i=i: abrir_nova_janela(f"Botão {i}"))

    root.mainloop()
criar_janela_principal()
